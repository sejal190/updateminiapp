import xml.etree.ElementTree as ET
import openpyxl
from pytest import *
from excel import *


def test_xml_to_excel():
    xml_file = 'Employee.xml'
    excel_file = 'Employee.xlsx'

    # Call the function
    xml_to_excel(xml_file, excel_file)

    # Load the excel file
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    # Check that the data was written to the sheet correctly
    headers = []
    for child in ET.parse(xml_file).getroot():
        for subchild in child:
            if subchild.tag not in headers:
                headers.append(subchild.tag)

    for i, header in enumerate(headers):
        assert sheet.cell(row=1, column=i + 1).value == header

    for i, child in enumerate(ET.parse(xml_file).getroot()):
        for j, subchild in enumerate(child):
            assert sheet.cell(row=i + 2, column=j + 1).value == subchild.text

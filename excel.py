import xml.etree.ElementTree as ET
import openpyxl
def xml_to_excel(xml_file, excel_file):
 # parse the xml file
 tree = ET.parse(xml_file)
 root = tree.getroot() # create a new excel file
 workbook = openpyxl.Workbook()
 sheet = workbook.active # get the list of column headers from the xml file
 headers = []
 for child in root:
  for subchild in child:
    if subchild.tag not in headers:
         headers.append(subchild.tag) # write the column headers to the first row of the excel sheet
 for i, header in enumerate(headers):
    sheet.cell(row=1, column=i + 1, value=header) # write the data from the xml file to the excel sheet
 for i, child in enumerate(root):
    for j, subchild in enumerate(child):
        sheet.cell(row=i + 2, column=j + 1, value=subchild.text) # save the excel file
 workbook.save(excel_file)# example usage
#xml_to_excel('Employee.xml', 'Employee.xlsx')